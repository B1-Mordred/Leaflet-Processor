from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    AnalyteDef,
    MeasurementRecord,
    RawCellRecord,
    WorkbookMeta,
    WorkbookParseResult,
)


_META_KEYS = {
    "order no.": "order_no",
    "lot no.": "lot_no",
    "exp. date": "exp_date",
    "consisting of": "consisting_of",
    "date of creation": "date_of_creation",
}


@dataclass(slots=True)
class _UnitBlockState:
    unit: str | None
    target_assigned: bool = False
    saw_range_low: bool = False
    range_high_assigned: bool = False


def parse_folder(folder: Path) -> list[WorkbookParseResult]:
    files = sorted(
        p for p in folder.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")
    )
    return [parse_workbook(path) for path in files]


def parse_workbook(path: Path) -> WorkbookParseResult:
    warnings: list[str] = []
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        raw_cells = _capture_raw_cells(workbook)
        rows_sheet = _find_rows_sheet(workbook.worksheets)
        if rows_sheet is None:
            warnings.append("No sheet matching 'sorted by rows' was found.")
            return WorkbookParseResult(
                source_file=path.name,
                workbook_meta=WorkbookMeta(sheet_name_rows=""),
                analytes=[],
                normalized_values=[],
                raw_cells=raw_cells,
                warnings=warnings,
            )

        workbook_meta = _extract_workbook_meta(rows_sheet, warnings)
        analytes, normalized_values = _extract_semantic_values(
            ws=rows_sheet, source_file=path.name, warnings=warnings
        )
        return WorkbookParseResult(
            source_file=path.name,
            workbook_meta=workbook_meta,
            analytes=analytes,
            normalized_values=normalized_values,
            raw_cells=raw_cells,
            warnings=warnings,
        )
    finally:
        workbook.close()


def _find_rows_sheet(worksheets: list[Worksheet]) -> Worksheet | None:
    for ws in worksheets:
        if "sorted by rows" in ws.title.lower():
            return ws
    return None


def _extract_workbook_meta(ws: Worksheet, warnings: list[str]) -> WorkbookMeta:
    key_rows: dict[str, int] = {}
    raw_values: dict[str, object] = {}
    for row_idx in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=1).value
        if not isinstance(key_cell, str):
            continue
        key_norm = key_cell.strip().lower()
        if key_norm in _META_KEYS:
            key_rows[key_norm] = row_idx
            raw_values[key_norm] = ws.cell(row=row_idx, column=2).value

    first_key_row = min(key_rows.values()) if key_rows else 1
    title_lines: list[str] = []
    for row_idx in range(1, first_key_row):
        value = ws.cell(row=row_idx, column=1).value
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                title_lines.append(cleaned)

    if "order no." not in key_rows:
        warnings.append("Metadata key 'Order No.' was not found.")
    if "substance" not in [str(ws.cell(r, 1).value).strip().lower() for r in range(1, 30)]:
        warnings.append("Could not locate 'Substance' row near the top of rows sheet.")

    return WorkbookMeta(
        title_lines=title_lines,
        order_no=_to_text(raw_values.get("order no.")),
        lot_no=_to_text(raw_values.get("lot no.")),
        exp_date=_to_date(raw_values.get("exp. date")),
        consisting_of=_to_text(raw_values.get("consisting of")),
        date_of_creation=_to_date(raw_values.get("date of creation")),
        sheet_name_rows=ws.title,
    )


def _extract_semantic_values(
    ws: Worksheet, source_file: str, warnings: list[str]
) -> tuple[list[AnalyteDef], list[MeasurementRecord]]:
    substance_row = _find_row_by_col_a(ws, "substance")
    if substance_row is None:
        warnings.append("No 'Substance' row found.")
        return [], []

    group_row = _find_row_by_col_a(ws, "group")
    analyte_cols: list[int] = []
    analyte_defs: list[AnalyteDef] = []
    for col_idx in range(4, ws.max_column + 1):
        raw_name = ws.cell(row=substance_row, column=col_idx).value
        name = _to_text(raw_name)
        if not name:
            continue
        group_name = _to_text(ws.cell(row=group_row, column=col_idx).value) if group_row else None
        analyte_cols.append(col_idx)
        analyte_defs.append(
            AnalyteDef(name=name, group_name=group_name, column_index=col_idx, units_seen=[])
        )

    if not analyte_defs:
        warnings.append("No analytes were discovered in 'Substance' row.")
        return [], []

    analyte_by_col = {a.column_index: a for a in analyte_defs}
    measurements: list[MeasurementRecord] = []
    current_block: _UnitBlockState | None = None
    measurement_started = False
    trailing_blank_rows = 0

    for row_idx in range(substance_row + 2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in analyte_cols]
        has_analyte_content = any(not _is_blank(v) for v in row_values)

        if not has_analyte_content:
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value
            c_val = ws.cell(row=row_idx, column=3).value
            if measurement_started and _is_blank(a_val) and _is_blank(b_val) and _is_blank(c_val):
                trailing_blank_rows += 1
                if trailing_blank_rows >= 8:
                    break
            continue

        measurement_started = True
        trailing_blank_rows = 0

        unit_text = _to_text(ws.cell(row=row_idx, column=3).value)
        if unit_text and (current_block is None or unit_text != current_block.unit):
            current_block = _UnitBlockState(unit=unit_text)
        effective_unit = unit_text if unit_text else (current_block.unit if current_block else None)
        if current_block is None:
            current_block = _UnitBlockState(unit=effective_unit)

        b_text = _to_text(ws.cell(row=row_idx, column=2).value)
        b_norm = b_text.lower() if b_text else ""
        is_separator_row = _row_is_separator(row_values)
        metric_role = _derive_metric_role(
            current_block=current_block,
            b_norm=b_norm,
            is_separator_row=is_separator_row,
        )

        sample_label = _to_text(ws.cell(row=row_idx, column=1).value)
        sample_code = b_text if b_norm != "range" else None
        if sample_code and sample_code == "-":
            sample_code = None

        for col_idx, raw_value in zip(analyte_cols, row_values):
            raw_text, numeric_value, value_status = _normalize_value(raw_value)
            analyte = analyte_by_col[col_idx]

            if effective_unit and value_status != "blank" and effective_unit not in analyte.units_seen:
                analyte.units_seen.append(effective_unit)

            measurements.append(
                MeasurementRecord(
                    source_file=source_file,
                    sample_label=sample_label,
                    sample_code=sample_code,
                    unit=effective_unit,
                    analyte_name=analyte.name,
                    group_name=analyte.group_name,
                    metric_role=metric_role,
                    raw_value=raw_text,
                    numeric_value=numeric_value,
                    value_status=value_status,
                    sheet_row=row_idx,
                    sheet_col=col_idx,
                )
            )

    if not measurements:
        warnings.append("No measurement rows were extracted.")
    return analyte_defs, measurements


def _derive_metric_role(
    current_block: _UnitBlockState, b_norm: str, is_separator_row: bool
) -> str:
    if b_norm == "range":
        current_block.saw_range_low = True
        return "range_low"

    if is_separator_row and current_block.saw_range_low and not current_block.range_high_assigned:
        return "range_sep"

    if is_separator_row:
        return "other"

    if not current_block.target_assigned:
        current_block.target_assigned = True
        return "target"

    if current_block.saw_range_low and not current_block.range_high_assigned:
        current_block.range_high_assigned = True
        return "range_high"

    return "other"


def _find_row_by_col_a(ws: Worksheet, key: str) -> int | None:
    key_norm = key.lower()
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if isinstance(value, str) and value.strip().lower() == key_norm:
            return row_idx
    return None


def _row_is_separator(values: list[object]) -> bool:
    non_blank = [v for v in values if not _is_blank(v)]
    if not non_blank:
        return False
    return all(_to_text(v) == "-" for v in non_blank)


def _normalize_value(value: object) -> tuple[str | None, float | None, str]:
    if _is_blank(value):
        return None, None, "blank"
    if isinstance(value, bool):
        return str(value), None, "text"
    if isinstance(value, (int, float)):
        return str(value), float(value), "ok"
    if isinstance(value, datetime):
        return value.isoformat(sep=" "), None, "text"
    if isinstance(value, date):
        return value.isoformat(), None, "text"

    text = str(value).strip()
    if not text:
        return None, None, "blank"
    lowered = text.lower()
    if lowered == "n.d.":
        return text, None, "nd"
    if text == "-":
        return text, None, "separator"

    candidate = text.replace(",", ".")
    try:
        parsed = float(candidate)
        return text, parsed, "ok"
    except ValueError:
        return text, None, "text"


def _capture_raw_cells(workbook) -> list[RawCellRecord]:
    records: list[RawCellRecord] = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                value = cell.value
                if _is_blank(value):
                    continue
                records.append(
                    RawCellRecord(
                        sheet_name=ws.title,
                        row_idx=cell.row,
                        col_idx=cell.column,
                        raw_value=_to_text(value),
                        python_type=type(value).__name__ if value is not None else None,
                    )
                )
    return records


def _to_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def _to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for pattern in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
    return None


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False

